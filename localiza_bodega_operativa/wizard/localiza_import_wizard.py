# -*- coding: utf-8 -*-
import base64
import io
from datetime import datetime, timedelta

from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None


class LocalizaImportWizard(models.TransientModel):
    _name = 'localiza.import.wizard'
    _description = 'Importador Excel Localiza'

    name = fields.Char(default='Importación Localiza')
    file = fields.Binary(string='Archivo Excel', required=True)
    filename = fields.Char(string='Nombre de archivo')
    import_type = fields.Selection([
        ('uniformes', 'Uniformes / Botas'),
        ('insumos', 'Insumos operativos'),
        ('gps', 'Inventario GPS'),
    ], required=True, default='insumos')
    location_id = fields.Many2one('stock.location', string='Ubicación de entrada', required=True,
                                  default=lambda self: self.env.ref('stock.stock_location_stock', raise_if_not_found=False))
    preview = fields.Text(string='Vista previa / resultado', readonly=True)
    create_stock = fields.Boolean(string='Crear stock inicial', default=True)

    def action_preview(self):
        rows = self._read_rows(limit=20)
        lines = []
        for row in rows[:20]:
            lines.append(' | '.join([str(x or '') for x in row[:12]]))
        self.preview = '\n'.join(lines) or _('No se encontraron filas legibles.')
        return self._reload()

    def action_import(self):
        if self.import_type == 'gps':
            count = self._import_gps()
        else:
            count = self._import_operational_products()
        self.preview = _('Importación completada. Registros procesados: %s') % count
        return self._reload()

    def _reload(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _workbook(self):
        if openpyxl is None:
            raise UserError(_('La librería Python openpyxl no está disponible en este servidor. Instálala en Odoo.sh o importa manualmente por CSV.'))
        data = base64.b64decode(self.file)
        return openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    def _read_rows(self, limit=None):
        wb = self._workbook()
        rows = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                vals = list(row)
                if any(v not in (None, '') for v in vals):
                    rows.append(vals)
                    if limit and len(rows) >= limit:
                        return rows
        return rows

    def _excel_date(self, value):
        if not value or value == 'X':
            return False
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, (int, float)):
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        try:
            return fields.Date.to_date(value)
        except Exception:
            return False

    def _clean(self, value):
        if value is None:
            return ''
        return str(value).strip()

    def _find_or_create_partner(self, name, supplier=False, customer=False):
        name = self._clean(name)
        if not name or name.upper() == 'X':
            return False
        partner = self.env['res.partner'].search([('name', '=ilike', name)], limit=1)
        if not partner:
            partner = self.env['res.partner'].create({'name': name, 'supplier_rank': 1 if supplier else 0, 'customer_rank': 1 if customer else 0})
        return partner

    def _find_or_create_product(self, name, default_code=None, tipo='insumo', tracking='none', cost=0.0):
        Product = self.env['product.product']
        domain = [('default_code', '=', default_code)] if default_code else [('name', '=ilike', name)]
        product = Product.search(domain, limit=1)
        if product:
            return product
        categ_xml = {
            'uniforme': 'product_category_localiza_uniformes',
            'bota': 'product_category_localiza_botas',
            'insumo': 'product_category_localiza_insumos',
            'gps': 'product_category_localiza_gps',
        }.get(tipo, 'product_category_localiza_insumos')
        categ = self.env.ref('localiza_bodega_operativa.%s' % categ_xml, raise_if_not_found=False)
        tmpl_vals = {
            'name': name,
            'default_code': default_code,
            'type': 'consu',
            'is_storable': True,
            'tracking': tracking,
            'standard_price': cost or 0.0,
            'x_localiza_tipo_operativo': tipo,
            'categ_id': categ.id if categ else False,
        }
        return Product.create(tmpl_vals)

    def _ensure_lot(self, product, serial):
        serial = self._clean(serial)
        if not serial or serial.upper() == 'X':
            return False
        lot = self.env['stock.lot'].search([('name', '=', serial), ('product_id', '=', product.id)], limit=1)
        if not lot:
            lot = self.env['stock.lot'].create({'name': serial, 'product_id': product.id, 'company_id': self.env.company.id})
        return lot

    def _update_stock(self, product, qty, lot=False):
        if not self.create_stock or qty <= 0:
            return
        self.env['stock.quant']._update_available_quantity(product, self.location_id, qty, lot_id=lot)

    def _import_operational_products(self):
        wb = self._workbook()
        count = 0
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            for r_index, row in enumerate(rows):
                vals = list(row)
                # Los archivos de Localiza tienen bloques repetidos de 9 columnas.
                for start in range(0, len(vals), 9):
                    block = vals[start:start + 9]
                    if len(block) < 9:
                        continue
                    header = [self._clean(x).upper() for x in block]
                    if 'CODIGO' in header or 'NO. PEDIDO' in header:
                        continue
                    item = self._clean(block[0])
                    code = self._clean(block[1])
                    if not item or not code or item.upper() in ('TALLA', 'INSUMO'):
                        continue
                    pedido = self._clean(block[2])
                    fecha = self._excel_date(block[3])
                    serie = self._clean(block[4])
                    factura = self._clean(block[5])
                    try:
                        price = float(block[6] or 0) if self._clean(block[6]).upper() != 'X' else 0.0
                    except Exception:
                        price = 0.0
                    proveedor = self._find_or_create_partner(block[7], supplier=True)
                    ubicacion = self._clean(block[8])
                    tipo = 'uniforme' if self.import_type == 'uniformes' else 'insumo'
                    product_name = item if self.import_type == 'insumos' else '%s %s' % (self._guess_block_title(rows, r_index, start), item)
                    tracking = 'serial' if serie and serie.upper() != 'X' else 'none'
                    product = self._find_or_create_product(product_name, code, tipo=tipo, tracking=tracking, cost=price)
                    product.product_tmpl_id.write({
                        'x_localiza_notas_operativas': 'Pedido: %s | Factura: %s | Ubicación origen: %s' % (pedido, factura, ubicacion),
                    })
                    lot = self._ensure_lot(product, serie) if tracking == 'serial' else False
                    self._update_stock(product, 1.0, lot=lot)
                    count += 1
        return count

    def _guess_block_title(self, rows, r_index, start):
        for back in range(r_index, max(-1, r_index - 4), -1):
            title = self._clean(rows[back][start] if start < len(rows[back]) else '')
            if title and title.upper() not in ('TALLA', 'INSUMO'):
                return title
        return 'Producto'

    def _import_gps(self):
        wb = self._workbook()
        count = 0
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            if len(vals) < 13:
                continue
            imei = self._clean(vals[2])
            modelo = self._clean(vals[3])
            marca = self._clean(vals[4])
            if not imei or imei.upper() == 'IMEI' or not modelo:
                continue
            proveedor = self._find_or_create_partner(vals[6], supplier=True)
            cliente = self._find_or_create_partner(vals[12], customer=True)
            try:
                price = float(vals[9] or 0)
            except Exception:
                price = 0.0
            product = self._find_or_create_product('GPS %s' % modelo, 'GPS-%s' % modelo, tipo='gps', tracking='serial', cost=price)
            lot = self._ensure_lot(product, imei)
            gps = self.env['localiza.gps.equipo'].search([('name', '=', imei)], limit=1)
            gps_vals = {
                'name': imei,
                'product_id': product.id,
                'lot_id': lot.id if lot else False,
                'modelo': modelo,
                'marca': marca,
                'fecha_compra': self._excel_date(vals[5]),
                'proveedor_id': proveedor.id if proveedor else False,
                'factura': self._clean(vals[7]),
                'no_serie': self._clean(vals[8]),
                'costo': price,
                'iva': float(vals[10] or 0) if vals[10] else 0.0,
                'placa': self._clean(vals[11]),
                'cliente_id': cliente.id if cliente else False,
                'state': 'instalado' if self._clean(vals[11]) or cliente else 'bodega',
            }
            if gps:
                gps.write(gps_vals)
            else:
                gps = self.env['localiza.gps.equipo'].create(gps_vals)
            self._update_stock(product, 1.0, lot=lot)
            count += 1
        return count
